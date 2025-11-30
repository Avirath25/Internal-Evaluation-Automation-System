# core/views.py
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt

from .models import ClassInfo, Student, Subject, Course, Marks

import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from math import ceil

# -------------------------
# PAGE RENDER VIEWS
# -------------------------
def login_page(request):
    return render(request, "index.html")

def landing_page(request):
    return render(request, "landing.html")

def teacher_home(request):
    return render(request, "teacher_home.html")

def select_class_page(request):
    return render(request, "selection.html")

def student_list_manage(request):
    return render(request, "student_list_manage.html")

def course_select_page(request):
    return render(request, "course_select.html")

def subject_create_page(request):
    return render(request, "subject_create.html")

def teacher_marks_page(request):
    return render(request, "marks_upload.html")

def student_login(request):
    return render(request, "student_login.html")

def student_summary_page(request):
    return render(request, "student_summary.html")

def course_and_upload_page(request):
    return render(request, "course_and_upload.html")

def student_dashboard_page(request):
    return render(request, "student_dashboard.html")

def student_login_page(request):
    return render(request, "student_login.html")

def teacher_login_page(request):
    return render(request, "teacher_login.html")

def four_credits_page(request):
    return render(request, "4credits.html")

def three_two_credits_page(request):
    return render(request, "3,2credits.html")

# -------------------------
# SUBJECT: CREATE + LIST
# -------------------------
@csrf_exempt
def create_subject(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"})

    data = json.loads(request.body)

    branch = data.get("branch")
    semester = data.get("semester")
    section = data.get("section")
    subject = data.get("subject")
    subcode = data.get("subcode")
    credits = data.get("credits")
    faculty = data.get("faculty", "")

    try:
        class_obj = ClassInfo.objects.get(branch=branch, semester=semester, section=section)
    except ClassInfo.DoesNotExist:
        return JsonResponse({"error": "Class not found"})

    new_sub = Subject.objects.create(
        class_info=class_obj,
        subject=subject,
        subcode=subcode,
        credits=credits,
        faculty=faculty
    )

    return JsonResponse({"status": "success", "subject_id": new_sub.id})


def list_subjects(request):
    branch = request.GET.get("branch")
    semester = request.GET.get("semester")
    section = request.GET.get("section")

    if not branch or not semester or not section:
        return JsonResponse({"subjects": [], "error": "Missing parameters"})

    try:
        class_obj = ClassInfo.objects.get(branch=branch, semester=semester, section=section)
    except ClassInfo.DoesNotExist:
        # Class doesn't exist yet, return empty list
        return JsonResponse({"subjects": [], "message": "Class not found. Please add students first."})

    subjects = Subject.objects.filter(class_info=class_obj).order_by('subject')

    out = []
    for s in subjects:
        out.append({
            "id": s.id,
            "subject": s.subject,
            "subcode": s.subcode,
            "credits": s.credits,
            "faculty": s.faculty
        })

    return JsonResponse({"subjects": out, "count": len(out)})


# -------------------------
# DOWNLOAD TEMPLATE
# -------------------------
def download_template_for_subject(request):
    """
    Returns an .xlsx template for the chosen subject.
    The generated template includes a "Total CIE (50)" column that contains an Excel formula
    matching the conversion logic:
      - For credits != 4: best-two of IA (out of 40) -> scaled to 25, assignments avg (out of 25) -> total = ceil(scaled25 + asgAvg)
      - For credits == 4: internal -> scaled to 15, assignments -> scaled to 10, lab_cie(15)+lab_test(10) -> final out of 50
    """
    subject_id = request.GET.get("subject_id")
    try:
        sub = Subject.objects.get(id=subject_id)
    except Subject.DoesNotExist:
        return JsonResponse({"error": "Subject not found"}, status=404)

    class_obj = sub.class_info

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    if int(sub.credits or 0) == 4:
        headers = ["SL No", "USN", "Name", "IA1", "IA2", "IA3",
                   "ASG1", "ASG2", "Lab CIE", "Lab Test", "Total CIE (50)"]
    else:
        headers = ["SL No", "USN", "Name", "IA1", "IA2", "IA3",
                   "ASG1", "ASG2", "Total CIE (50)"]
    ws.append(headers)

    # style header
    header_font = Font(bold=True)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fetch students for class
    students = Student.objects.filter(class_info=class_obj).order_by("sl_no")

    # Build formula template depending on credits
    if int(sub.credits or 0) == 4:
        # 4-credit: IA(15) + Asg(10) + Lab(25) = 50 (columns D-J, total in K)
        formula_template = (
            "=CEILING( ( (LARGE(D{r}:F{r},1)+LARGE(D{r}:F{r},2))/2 )/40*15"
            " + ( (G{r}+H{r})/2 )/25*10"
            " + (I{r}+J{r}), 1)"
        )
    else:
        # 3/2/1-credit: IA(25) + Asg(25) = 50 (columns D-H, total in I)
        formula_template = (
            "=CEILING( ( (LARGE(D{r}:F{r},1)+LARGE(D{r}:F{r},2))/2 )/40*25"
            " + (G{r}+H{r})/50*25, 1)"
        )

    # write rows
    for idx, s in enumerate(students, start=2):
        ws.cell(row=idx, column=1, value=s.sl_no)
        ws.cell(row=idx, column=2, value=s.usn)
        name_cell = ws.cell(row=idx, column=3, value=s.name)
        name_cell.alignment = Alignment(wrap_text=True)
        if int(sub.credits or 0) == 4:
            # blank cells for marks D..J (4-credit has lab columns)
            for col in range(4, 11):
                ws.cell(row=idx, column=col, value="")
            # put formula in Total column (column 11)
            ws.cell(row=idx, column=11, value=formula_template.format(r=idx))
        else:
            # blank cells for marks D..H (3/2/1-credit no lab columns)
            for col in range(4, 9):
                ws.cell(row=idx, column=col, value="")
            # put formula in Total column (column 9)
            ws.cell(row=idx, column=9, value=formula_template.format(r=idx))

    # adjust column widths (wider for name)
    if int(sub.credits or 0) == 4:
        col_widths = {
            1: 6,   # SL No
            2: 12,  # USN
            3: 28,  # Name
            4: 8,  # IA1
            5: 8,  # IA2
            6: 8,  # IA3
            7: 8,  # ASG1
            8: 8,  # ASG2
            9: 10, # Lab CIE
            10: 10,# Lab Test
            11: 12 # Total
        }
    else:
        col_widths = {
            1: 6,   # SL No
            2: 12,  # USN
            3: 28,  # Name
            4: 8,  # IA1
            5: 8,  # IA2
            6: 8,  # IA3
            7: 8,  # ASG1
            8: 8,  # ASG2
            9: 12  # Total
        }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    file_name = f"{sub.subject}_template.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    wb.save(response)
    return response


# -------------------------
# UPLOAD MARKS
# -------------------------
@csrf_exempt
def upload_marks_for_subject(request):
    """
    Expects form-data:
      - file: .xlsx file
      - subject_id: id of Subject (from Subject model)
      - teacher_name (optional)
      - credits (optional)
    Behavior:
      - find Subject by id
      - find or create Course for the same class_info with the subject name (so Marks.course can reference it)
      - parse sheet rows, calculate total (if missing) and save into Marks (student, class_info, course)
    """
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=400)

    file = request.FILES.get("file")
    subject_id = request.POST.get("subject_id")
    teacher_name = request.POST.get("teacher_name", "").strip()
    credits_override = request.POST.get("credits")

    if not file or not subject_id:
        return JsonResponse({"error": "file and subject_id required"}, status=400)

    subject = get_object_or_404(Subject, id=subject_id)
    class_obj = subject.class_info

    # create/find a Course instance so Marks.course can point to it
    course_obj, _ = Course.objects.get_or_create(
        class_info=class_obj,
        course_name=subject.subject,
        sub_code=getattr(subject, "subcode", "") or "",
        defaults={"credits": subject.credits or (int(credits_override) if credits_override else 3),
                  "faculty": subject.faculty or teacher_name}
    )

    # if teacher_name provided, update course faculty; if credits_override provided, update course credits
    updated = False
    if teacher_name and course_obj.faculty != teacher_name:
        course_obj.faculty = teacher_name
        updated = True
    if credits_override:
        try:
            cval = int(credits_override)
            if course_obj.credits != cval:
                course_obj.credits = cval
                updated = True
        except:
            pass
    if updated:
        course_obj.save()

    # load workbook
    try:
        wb = openpyxl.load_workbook(file, data_only=False)
    except Exception as e:
        return JsonResponse({"error": f"Failed to read Excel file: {str(e)}"}, status=400)

    sheet = wb.active

    # build header mapping
    headers = []
    try:
        headers = [str(h.value).lower().strip() if h.value is not None else "" for h in sheet[1]]
    except Exception:
        return JsonResponse({"error": "Invalid sheet header row"}, status=400)

    def get_index(names):
        for name in names:
            try:
                if name in headers:
                    return headers.index(name)
            except Exception:
                continue
        return None

    i_usn = get_index(["usn", "u s n", "usn "])
    i_ia1 = get_index(["ia1 (40)", "ia1", "ia 1", "ia1 ( 40 )"])
    i_ia2 = get_index(["ia2 (40)", "ia2", "ia 2"])
    i_ia3 = get_index(["ia3 (40)", "ia3", "ia 3"])
    i_asg1 = get_index(["asg1 (25)", "asg1 (20)", "asg1", "asg 1"])
    i_asg2 = get_index(["asg2 (25)", "asg2 (20)", "asg2", "asg 2"])
    i_labcie = get_index(["lab cie", "lab cie (15)", "lab_cie"])
    i_labtest = get_index(["lab test", "lab test (10)", "lab_test"])
    i_total = get_index(["total cie", "total cie (50)", "total", "total cie ( 50 )"])

    # row loop
    saved = 0
    errors = []
    credits_for_calc = course_obj.credits or subject.credits or (int(credits_override) if credits_override else 3)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # read usn and skip empty rows
        usn = None
        if i_usn is not None and i_usn < len(row):
            usn = row[i_usn]
        # if usn not provided, try column 2 as fallback
        if not usn and len(row) >= 2:
            usn = row[1]

        if not usn:
            continue
        usn = str(usn).strip()

        # find student in that class
        try:
            student = Student.objects.get(usn=usn, class_info=class_obj)
        except Student.DoesNotExist:
            # skip rows whose USN not found (optionally collect)
            errors.append(f"Student with USN {usn} not found in class {class_obj}.")
            continue

        def val(i):
            try:
                if i is None: 
                    return None
                if i >= len(row): 
                    return None
                v = row[i]
                if v in ("", None):
                    return None
                return float(v)
            except Exception:
                return None

        v_ia1 = val(i_ia1)
        v_ia2 = val(i_ia2)
        v_ia3 = val(i_ia3)
        v_asg1 = val(i_asg1)
        v_asg2 = val(i_asg2)
        v_labcie = val(i_labcie) or 0
        v_labtest = val(i_labtest) or 0
        v_total = val(i_total)

        # If total is not provided in sheet, compute it using same logic as template formulas
        if v_total is None:
            # prepare internals list
            internals = [x for x in (v_ia1, v_ia2, v_ia3) if x is not None]
            # need at least two internals to compute best-two avg; if less, treat missing as 0 (you may change)
            if len(internals) < 2:
                # set missing internals to 0 if needed to compute
                while len(internals) < 2:
                    internals.append(0.0)
            # sort descending and pick top2
            internals_sorted = sorted(internals, reverse=True)
            best2_avg = (internals_sorted[0] + internals_sorted[1]) / 2.0

            asg_avg = None
            if v_asg1 is not None and v_asg2 is not None:
                asg_avg = (v_asg1 + v_asg2) / 2.0
            elif v_asg1 is not None:
                asg_avg = v_asg1
            elif v_asg2 is not None:
                asg_avg = v_asg2
            else:
                asg_avg = 0.0

            if int(credits_for_calc) == 4:
                # 4-credit: IA(15) + Asg(10) + Lab(25) = 50
                internal15 = (best2_avg / 40.0) * 15.0
                assign10 = (asg_avg / 25.0) * 10.0
                lab_total = (v_labcie or 0) + (v_labtest or 0)
                total_calc = internal15 + assign10 + lab_total
                v_total = ceil(total_calc)
            else:
                # 3/2/1-credit: IA(25) + Asg(25) = 50
                internal25 = (best2_avg / 40.0) * 25.0
                assign_sum = (v_asg1 or 0) + (v_asg2 or 0)
                assign25 = (assign_sum / 50.0) * 25.0  # Scale from 50 to 25
                total_calc = internal25 + assign25
                v_total = ceil(total_calc)
        # else: if total provided, accept as float
        else:
            try:
                v_total = float(v_total)
            except:
                v_total = None

        # Save/Update Marks object for this student + course
        # Marks has fields: student, class_info, course, ia1, ia2, ia3, asg1, asg2, lab_cie, lab_test, total
        try:
            obj, created = Marks.objects.update_or_create(
                student=student,
                class_info=class_obj,
                course=course_obj,
                defaults={
                    "ia1": v_ia1,
                    "ia2": v_ia2,
                    "ia3": v_ia3,
                    "asg1": v_asg1,
                    "asg2": v_asg2,
                    "lab_cie": v_labcie,
                    "lab_test": v_labtest,
                    "total": v_total
                }
            )
            saved += 1
        except Exception as e:
            errors.append(f"Failed saving marks for {usn}: {str(e)}")

    result = {"status": "saved", "saved_rows": saved}
    if errors:
        result["errors"] = errors[:20]  # don't return too big list

    return JsonResponse(result)


# -------------------------
# STUDENT LIST
# -------------------------
def get_student_list(request):
    branch = request.GET.get("branch")
    semester = request.GET.get("semester")
    section = request.GET.get("section")

    try:
        class_obj = ClassInfo.objects.get(branch=branch, semester=semester, section=section)
    except ClassInfo.DoesNotExist:
        return JsonResponse({"students": []})

    students = [
        {"sl": s.sl_no, "usn": s.usn, "name": s.name}
        for s in class_obj.students.all().order_by("sl_no")
    ]

    return JsonResponse({"students": students})


# -------------------------
# UPLOAD STUDENT LIST
# -------------------------
@csrf_exempt
def upload_student_list(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=400)

    data = json.loads(request.body)
    branch = data["branch"]
    semester = data["semester"]
    section = data["section"]
    students = data["students"]

    class_obj, _ = ClassInfo.objects.get_or_create(
        branch=branch,
        semester=semester,
        section=section
    )

    Student.objects.filter(class_info=class_obj).delete()

    for s in students:
        Student.objects.create(
            class_info=class_obj,
            sl_no=s["sl"],
            usn=s["usn"],
            name=s["name"]
        )

    return JsonResponse({"status": "success"})


# -------------------------
# STUDENT SUMMARY
# -------------------------
def get_student_summary(request):
    usn = request.GET.get("usn")
    subject_id = request.GET.get("subject_id")

    if not usn or not subject_id:
        return JsonResponse({"error": "USN and subject_id required"}, status=400)

    try:
        stu = Student.objects.get(usn=usn)
    except Student.DoesNotExist:
        return JsonResponse({"error": "Student not found"}, status=404)

    try:
        marks = Marks.objects.get(student=stu, course__id=subject_id)
    except Marks.DoesNotExist:
        return JsonResponse({"error": "Marks not found for this subject"}, status=404)

    course = marks.course

    return JsonResponse({
        "status": "success",
        "summary": {
            "name": stu.name,
            "usn": stu.usn,
            "semester": stu.class_info.semester,
            "section": stu.class_info.section,
            "subject": course.course_name,
            "sub_code": course.sub_code,
            "faculty": course.faculty,
            "credits": course.credits,

            "ia1": marks.ia1,
            "ia2": marks.ia2,
            "ia3": marks.ia3,
            "asg1": marks.asg1,
            "asg2": marks.asg2,
            "lab_cie": marks.lab_cie,
            "lab_test": marks.lab_test,
            "total": marks.total
        }
    })

# -------------------------
# STUDENT LOGIN CHECK
# -------------------------
@csrf_exempt
def student_check(request):
    data = json.loads(request.body)
    usn = data.get("usn")

    student = Student.objects.filter(usn=usn).first()
    if not student:
        return JsonResponse({"error": "USN not found"}, status=404)

    return JsonResponse({"status": "ok", "student_name": student.name})
def student_subjects(request):
    usn = request.GET.get("usn")
    if not usn:
        return JsonResponse({"status": "error", "error": "USN required"}, status=400)

    student = Student.objects.filter(usn=usn).first()
    if not student:
        return JsonResponse({"status": "error", "error": "Student not found"}, status=404)

    marks = Marks.objects.filter(student=student)

    subject_list = []
    for m in marks:
        subject_list.append({
            "subject_id": m.course.id if m.course else None,
            "subject": m.course.course_name if m.course else "",
            "subcode": m.course.sub_code if m.course else "",
            "credits": m.course.credits if m.course else "",
            "faculty": m.course.faculty if m.course else "",
            "has_marks": True
        })

    return JsonResponse({"status": "success", "subjects": subject_list})


# -------------------------
# GET UPLOADED MARKS LIST
# -------------------------
def get_uploaded_marks(request):
    """Get list of subjects with uploaded marks for a specific class"""
    branch = request.GET.get("branch")
    semester = request.GET.get("semester")
    section = request.GET.get("section")

    if not branch or not semester or not section:
        return JsonResponse({"uploads": []})

    try:
        class_obj = ClassInfo.objects.get(branch=branch, semester=semester, section=section)
    except ClassInfo.DoesNotExist:
        return JsonResponse({"uploads": []})

    # Get all courses for this specific class that have marks
    courses = Course.objects.filter(class_info=class_obj)
    
    uploads = []
    for course in courses:
        # Count marks for this course in this specific class
        student_count = Marks.objects.filter(
            course=course,
            class_info=class_obj
        ).count()
        
        # Only include if there are marks uploaded
        if student_count > 0:
            uploads.append({
                "course_id": course.id,
                "subject": course.course_name,
                "subcode": course.sub_code or "",
                "credits": course.credits,
                "faculty": course.faculty or "",
                "student_count": student_count
            })

    return JsonResponse({"uploads": uploads})


# -------------------------
# GET MARKS FOR A COURSE
# -------------------------
def get_course_marks(request):
    """Get all marks data for a specific course"""
    course_id = request.GET.get("course_id")

    if not course_id:
        return JsonResponse({"error": "course_id required"}, status=400)

    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return JsonResponse({"error": "Course not found"}, status=404)

    # Get all marks for this course
    marks = Marks.objects.filter(course=course).select_related('student').order_by('student__sl_no')

    marks_list = []
    for m in marks:
        marks_list.append({
            "usn": m.student.usn,
            "name": m.student.name,
            "ia1": m.ia1,
            "ia2": m.ia2,
            "ia3": m.ia3,
            "asg1": m.asg1,
            "asg2": m.asg2,
            "lab_cie": m.lab_cie,
            "lab_test": m.lab_test,
            "total": m.total,
            "credits": course.credits
        })

    return JsonResponse({
        "status": "success",
        "marks": marks_list,
        "course_name": course.course_name,
        "credits": course.credits
    })


# -------------------------
# DELETE UPLOADED MARKS
# -------------------------
@csrf_exempt
def delete_uploaded_marks(request):
    """Delete all marks for a specific course (subject)"""
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=400)

    data = json.loads(request.body)
    course_id = data.get("course_id")

    if not course_id:
        return JsonResponse({"error": "course_id required"}, status=400)

    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return JsonResponse({"error": "Course not found"}, status=404)

    # Delete all marks for this course
    deleted_count = Marks.objects.filter(course=course).delete()[0]

    return JsonResponse({
        "status": "success",
        "message": f"Deleted marks for {deleted_count} students",
        "deleted_count": deleted_count
    })
